import os
import csv
import xlrd
import openpyxl
from reportlab.lib.pagesizes import letter
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib import colors
import sys
def convert_excel_to_pdf(input_filename, pdf_filename):
    _, extension = os.path.splitext(input_filename)
    
    if extension.lower() == ".csv":
        data = []
        with open(input_filename, "r") as csvfile:
            csvreader = csv.reader(csvfile)
            for row in csvreader:
                data.append(row)
    elif extension.lower() == ".xls":
        workbook = xlrd.open_workbook(input_filename)
        worksheet = workbook.sheet_by_index(0)
        
        data = []
        for row_index in range(worksheet.nrows):
            row_data = []
            for col_index in range(worksheet.ncols):
                row_data.append(worksheet.cell_value(row_index, col_index))
            data.append(row_data)
    elif extension.lower() == ".xlsx":
        workbook = openpyxl.load_workbook(input_filename)
        worksheet = workbook.active
        
        data = []
        for row in worksheet.iter_rows(values_only=True):
            data.append(list(row))
    else:
        raise ValueError("Unsupported file format")
    
    # Create a PDF using ReportLab
    doc = SimpleDocTemplate(pdf_filename, pagesize=letter)
    elements = []
    
    table_data = []
    for row in data:
        table_row = []
        for cell_content in row:
            if isinstance(cell_content, float):
                cell_content = str(cell_content)  # Convert float to string
            p = Paragraph(cell_content, getSampleStyleSheet()["Normal"])
            table_row.append(p)
        table_data.append(table_row)
    
    table = Table(table_data, colWidths=[60] * len(data[0]))  # Adjust colWidths as needed
    style = TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
    ])
    table.setStyle(style)
    
    elements.append(table)
    doc.build(elements)

if __name__ == "__main__":
    input_filename = sys.argv[1]  # Replace with your file name
    pdf_filename = sys.argv[2]       # Output PDF filename
    dir_path, filename = os.path.split(input_filename)
    base_name, extension = os.path.splitext(filename)
    new_base_name = f"{base_name}_converted.pdf"
    new_path = os.path.join(dir_path, f"{new_base_name}")
    _, extension = os.path.splitext(input_filename)
    pdf_filename = new_path
    if extension.lower() == ".xlsx":
        # Convert XLSX to temporary CSV
        temp_csv_filename = "temp.csv"
        try:
            workbook = openpyxl.load_workbook(input_filename)
            sheet = workbook.active
            with open(temp_csv_filename, "w", newline="") as csvfile:
                csvwriter = csv.writer(csvfile)
                for row in sheet.iter_rows():
                    csvwriter.writerow([cell.value for cell in row])
            
            # Call convert_excel_to_pdf with the temporary CSV file
            convert_excel_to_pdf(temp_csv_filename, pdf_filename)
            print("XLSX file converted to PDF successfully using ReportLab!")

        except Exception as e:
            print(f"Error converting XLSX to CSV and PDF: {e}")
        finally:
            # Delete the temporary CSV file
            if os.path.exists(temp_csv_filename):
                os.remove(temp_csv_filename)
    else:
        # Call convert_excel_to_pdf directly for other file formats
        convert_excel_to_pdf(input_filename, pdf_filename)
        print("File converted to PDF successfully using ReportLab!")
